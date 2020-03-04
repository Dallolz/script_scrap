import scrapy
import logging
import sys
sys.path.append("..")
import json
import time
import os, sys
from openpyxl import load_workbook
from ..items import ScriptV2Item
from scrapy.selector import Selector
from scrapy.utils.project import get_project_settings
from urlparse import urlparse, parse_qs
from twisted.internet import reactor
from scrapy.crawler import CrawlerRunner
from scrapy.utils.log import configure_logging

logger = logging.getLogger('mycustomlogger')

class AddonSimilarSpider(scrapy.Spider):
    name = "addonsimilartech"
    start_urls = []
    count = -1

    def __init__(self):
        self.start_urls = []
        wb = load_workbook(filename = 'Hunter.xlsx', data_only=True)
        worksheet = wb.get_sheet_by_name('Feuil1')
        for row_cells in worksheet.iter_rows():
          for cell in row_cells:
            self.start_urls.append('https://addon.similartech.com/addons/a/0.11.4/chrome/79.0.3945.79/discover?url=%s' % cell.value)


    def parse(self, response):
        global m_visits
        global c_name
        m_visits = "N/D"
        c_name = "N/D"
        try:
	        json_data = json.loads(response.body.split('window.initialData.discover = ')[1].strip().split('</script>')[0].strip().replace('};','}').strip())
	        info = json_data['discoverData']['info']
	        m_visits = info['monthlyVisits']
	        c_name = info['countryName']
	        print ("monthly visits: %s", monthly_visits)
	        print ("country name: %s", country_name)

        except:
	        pass
        return ScriptV2Item(monthly_visits=m_visits, country_name= c_name, domain_list=response.request.url)
