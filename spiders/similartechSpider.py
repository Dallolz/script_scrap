import scrapy
import sys
sys.path.append("..")
import logging
import json
import time
import os, sys
from openpyxl import load_workbook
from ..items import ScriptV1Item
from scrapy.selector import Selector
from scrapy.utils.project import get_project_settings
from urlparse import urlparse, parse_qs
from twisted.internet import reactor
from scrapy.crawler import CrawlerRunner
from scrapy.utils.log import configure_logging

sys.path.append("..")
logger = logging.getLogger('mycustomlogger')

class SimilarTechSpider(scrapy.Spider):
    name = "similartech"
    start_urls = []
    count = -1

    def __init__(self):
        self.start_urls = []
        wb = load_workbook(filename = 'Hunter.xlsx', data_only=True)
        worksheet = wb.get_sheet_by_name('Feuil1')
        for row_cells in worksheet.iter_rows():
          for cell in row_cells:
            self.start_urls.append('https://www.similartech.com/websites/%s' % cell.value)


    def parse(self, response):
        global pname
        global cname
        pname = "N/D"
        cname = "N/D"
        sel = Selector(response)
        sections  = sel.xpath('//div[@class="row cmp "]')
        for s in sections:
	        if "ECommerce" == "".join(s.xpath('div[1]/div[@class="row cmp-header sub"]/a/h2/text()').extract()).strip():
		        sub_sections = s.xpath('div[1]/div[@class="row sub"]/div[1]/div[@class="row cmp sub-category"]')
		        for sub in sub_sections:
			        if "ECommerce Platforms" == "".join(sub.xpath('div[1]/div[@class="row cmp-header sub"]/a/h2/text()').extract()).strip():
				        platforms = sub.xpath('div[1]/div[@class="row sub"]/div[1]/div[@class="row list-item-block"]')
				        for p in platforms:
					        pname = "".join(p.xpath('div[2]/div[@class="item-title"]/a[@class="tech-name"]/text()').extract()).strip()
					        break
	        elif "Customer Communication" == "".join(s.xpath('div[1]/div[@class="row cmp-header sub"]/a/h2/text()').extract()).strip():
		        services = s.xpath('div[1]/div[@class="row sub"]/div[1]/div[@class="row list-item-block"]')
		        for s in services:
			        cname = "".join(s.xpath('div[2]/div[@class="item-title"]/a[@class="tech-name"]/text()').extract()).strip()
			        break
        return ScriptV1Item(customer_service_title=cname, commerce_platform_title=pname, domain_list=response.request.url)


# def get_all():


# #### main 
# def parse_sheet():
#   ## recup les domains d'excel
#   return ["1.com", "2.com"]

# def crawl(domain_list)
#   for domain in domain_list:
#     ## spider addon
#     ## spider similartech
#     ## cf la doc pour lancer plusieurs spider en meme temps: process.start()
#     ## recuperer dict avec les infos
#     ## write_line(info)

# def write_line(info):
#   ## cell 1 ecrit info.get("domain")
#   ## cell 2 ecrit info.get("stack")
#   ## cell 3 ecrit info.get("leading_country")
#   ## cell 4 ecrit info.get("monthly_view")

# def write_xlsx(infos):
#   for info in infos:
#     write_line(info)

# def get_all(domains):
#   infos = []
#   for domain in domains:
#     info = get_info(domain)
#     infos.append(info)
#   return infos

# def get_info(domain):
#   stats = get_stats(domain)
#   stack = get_stack(domain)
#   return zip({domain: domain}, stats, stack)

# def get_website(url):
#   ## recup page web
#   ## return html
#   return "string2html"

# def get_stack(domain):
#   ## build url pour similartech
#   url = "simliartech.com/webiste/" + domain  
#   html = get_website(url)
#   ## parsing special pour recup la stack sous la div avec un selecteur specifique au portail similartech
#   return {stack: "stack1"}

# def get_stats(domain):
#     ## build url pour similartech
#   url = "simliartech.com/webiste/" + domain  
#   ## requete vers addon.hgdjshgd
#   html = get_website(domain)
#   ## parsing specailpour recup des stats specifique au site addon de similartech
#   return {
#     "leading country": "fr",
#      "monthly_views": 12000
#      }


# class ChromeSimilarSpider(scrapy.Spider):
#   name = "ChromeStech"
#   def parse(self, response):
#         for qulaote in response.css('div.quote'):
#             yield {
#               'text': quote.css('span.text::text').get(),
#               'author': quote.css('small.author::text').get(),
#               'tags': quote.css('div.tags a.tag::text').getall(),
#             }
#         filename = response.url.split("/")[-2]
#         open(filename, 'wb').write(your_data)

# process = CrawlerProcess()
# process.crawl(SimilarTechSpider)
# process.crawl(ChromeSimilarSpider)
# process.crawl(WebDomainsSpider)
# process.start()